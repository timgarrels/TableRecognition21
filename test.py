import logging
import random
import sys
from os.path import join

from openpyxl import load_workbook

from LabelRegionPreprocessor import LabelRegionPreprocessor
from SpreadSheetGraph import SpreadSheetGraph

logging.basicConfig(stream=sys.stdout, level=logging.INFO)

logger = logging.getLogger(__name__)

DATA_DIR = "data"
VISUALIZATIONS_DIR = "visualizations"
ANNOTATIONS = "annotations_elements.json"
SPREADSHEET = "andrea_ring__3__HHmonthlyavg.xlsx"
ANNOTATION_FILE = join(DATA_DIR, ANNOTATIONS)
SPREADSHEET_FILE = join(DATA_DIR, SPREADSHEET)


def main():
    logger.info("Creating Label Regions...")
    wb = load_workbook(SPREADSHEET_FILE)
    sheetname = wb.sheetnames[0]
    sheet = wb[sheetname]
    preprocessor = LabelRegionPreprocessor()

    label_regions = preprocessor.preproces_annotations(
        ANNOTATION_FILE,
        SPREADSHEET_FILE,
        sheetname,
    )

    sheet_graph = SpreadSheetGraph.from_label_regions_and_sheet(label_regions, sheet)
    logger.debug(f"Edge Count: {len(sheet_graph.edge_list)}")

    sheet_graph.edge_toggle_list = [random.choice([True, False]) for _ in range(len(sheet_graph.edge_toggle_list))]
    sheet_graph.visualize(out=join(VISUALIZATIONS_DIR, 'test_graph_original'))
    components = sheet_graph.get_components()
    for component in components:
        print(component.header_groups)


if __name__ == "__main__":
    main()
