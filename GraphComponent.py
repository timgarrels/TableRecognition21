"""Class to represent a Graph Component, used to get attributes/data and cache them"""

from typing import List

from openpyxl.utils.cell import get_column_letter

from SpreadSheetGraph import SpreadSheetGraph


class GraphComponent(object):
    def __init__(self, graph: SpreadSheetGraph, component: List[int]):
        self.graph = graph
        self.label_regions = component

    @property
    def bounding_box(self):
        """Returns the minimum bounding box of this component"""
        if self._bounding_box is None:
            top_lefts = [self.graph.nodes[i]["top_left"] for i in self.label_regions]
            bottom_rights = [self.graph.nodes[i]["bottom_right"] for i in self.label_regions]

            min_xs = [top_left[0] for top_left in top_lefts]
            min_ys = [top_left[1] for top_left in top_lefts]
            max_xs = [bottom_right[0] for bottom_right in bottom_rights]
            max_ys = [bottom_right[1] for bottom_right in bottom_rights]

            min_x = min(min_xs)
            min_y = min(min_ys)
            max_x = max(max_xs)
            max_y = max(max_ys)
            self._bounding_box = {"top_left": (min_x, min_y), "bottom_right": (max_x, max_y)}
        return self._bounding_box

    @property
    def data(self):
        """Returns all data label regions"""
        if self._data is None:
            self._data = [i for i in self.label_regions if self.graph.nodes[i]["type"] == "Data"]
        return self._data

    @property
    def heads(self):
        """Returns all header label regions"""
        if self._heads is None:
            self._heads = [i for i in self.label_regions if self.graph.nodes[i]["type"] == "Header"]
        return self._heads

    def area(self, label_region_index):
        """Returns the number of cells in the given label region"""
        if self._area is None:
            self._area = {}
        if self._area.get(label_region_index, None) is None:
            label_region = self.graph.nodes[label_region_index]
            min_x, min_y = label_region["top_left"]
            max_x, max_y = label_region["bottom_right"]
            x = len(range(min_x, max_x + 1))
            y = len(range(min_y, max_y + 1))
            self._area[label_region_index] = x * y
        return self._area[label_region_index]

    def rows(self, label_region_index):
        """Returns the row indices wihtin the given label region"""
        if self._rows is None:
            self._rows = {}
        if self._rows.get(label_region_index, None) is None:
            label_region = self.graph.nodes[label_region_index]
            min_x, min_y = label_region["top_left"]
            max_x, max_y = label_region["bottom_right"]
            self._rows[label_region_index] = list(range(min_x, max_x + 1))
        return self._rows[label_region_index]

    def cols(self, label_region_index):
        """Returns the column indices wihtin the given label region"""
        if self._cols is None:
            self._cols = {}
        if self._cols.get(label_region_index, None) is None:
            label_region = self.graph.nodes[label_region_index]
            min_x, min_y = label_region["top_left"]
            max_x, max_y = label_region["bottom_right"]
            self._cols[label_region_index] = list(range(min_y, max_y + 1))
        return self._cols[label_region_index]

    def cwidth(self, c_index):
        """Returns the width of the coloumn :c_index"""
        return self.graph.sheet.column_dimensions[get_column_letter(c_index)].width

    def rheight(self, r_index):
        """Returns the height of the row :r_index"""
        return self.graph.sheet.row_dimensions[r_index].height

    @property
    def get_header_groups(self):
        if self._header_groups is None:
            # TODO: Merge and list header groups
            self._header_groups = 1
        return self._header_groups

    @property
    def avg_waec(self):
        # Assumption: "adjacent empty columns" are just all empty columns in our partition
        bounds = self.bounding_box
        min_x = bounds["top_left"][0]
        max_x = bounds["bottom_right"][0]
        empty_cols = []
        for col in self.graph.ws.iter_cols(min_col=min_x, max_col=max_x, values_only=True):
