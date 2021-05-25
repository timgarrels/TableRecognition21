import logging
import sys
from os import getcwd
from os.path import join

from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet

from graph.GraphComponentData import GraphComponentData
from graph.SpreadSheetGraph import SpreadSheetGraph
from labelregions.LabelRegionPreprocessor import LabelRegionPreprocessor
from rater.Rater import Rater
from search.ExhaustiveSearch import ExhaustiveSearch
from search.GeneticSearch import GeneticSearch
from search.GeneticSearchConfiguration import GeneticSearchConfiguration
from visualization.GraphVisualization import visualize_graph
from visualization.LabelRegionVisualization import visualize_lrs
from visualization.TableDefinitionVisualization import visualize_table_definition

logging.basicConfig(stream=sys.stdout, level=logging.INFO)

logger = logging.getLogger(__name__)
logger.setLevel(logging.DEBUG)

DATA_DIR = join(getcwd(), "data")
VISUALIZATIONS_DIR = join(getcwd(), "visualizations")
ANNOTATIONS = "annotations_elements.json"
SPREADSHEET = "andrea_ring__3__HHmonthlyavg.xlsx"
ANNOTATION_FILE = join(DATA_DIR, ANNOTATIONS)
SPREADSHEET_FILE = join(DATA_DIR, SPREADSHEET)


def main():
    logger.info("Creating Label Regions...")
    wb = load_workbook(SPREADSHEET_FILE)
    sheetname = wb.sheetnames[0]
    sheet: Worksheet = wb[sheetname]
    preprocessor = LabelRegionPreprocessor()

    label_regions = preprocessor.preproces_annotations(
        ANNOTATION_FILE,
        SPREADSHEET_FILE,
        sheetname,
    )

    visualize_lrs(label_regions, out=join(VISUALIZATIONS_DIR, 'lrs.xlsx'))

    sheet_graph = SpreadSheetGraph.from_label_regions_and_sheet(label_regions, sheet)
    logger.debug(f"Edge Count: {len(sheet_graph.edge_list)}")

    visualize_graph(sheet_graph, out=join(VISUALIZATIONS_DIR, 'original_graph'))

    if len(sheet_graph.nodes) <= 10:
        # Less than 11 nodes, do exhaustive search
        search = ExhaustiveSearch(
            sheet_graph,
            Rater(),
        )
    else:
        search = GeneticSearch(
            sheet_graph,
            # TODO: Rater should be trained
            Rater(),
            GeneticSearchConfiguration(),
        )

    fittest, fittest_rating = search.run()

    logger.debug(f"Best rating: {fittest_rating}")
    logger.debug(f"Best individual: {fittest}")

    logger.info("Visualizing Fittest Graph Partition...")
    sheet_graph.edge_toggle_list = fittest

    table_definitions = [GraphComponentData(component, sheet_graph).bounding_box for component in
                         sheet_graph.get_components()]

    visualize_graph(sheet_graph, out=join(VISUALIZATIONS_DIR, 'fittest_graph'))
    visualize_table_definition(label_regions, table_definitions, out=join(VISUALIZATIONS_DIR, 'lrs.xlsx'))


if __name__ == "__main__":
    main()
