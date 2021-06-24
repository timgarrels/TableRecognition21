import json
import logging
from functools import cached_property
from os import listdir
from os.path import isfile, join, split, getsize
from typing import Generator, List

from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet

from labelregions.LabelRegionLoader import LabelRegionLoader
from loader.SheetData import SheetData

logger = logging.getLogger(__name__)
logger.setLevel(logging.INFO)


# TODO: Preprocess Sheets to remove all hidden rows and columns
#       The Chair Annotations are based on csv's, which do not contain hidden cols/rows. This means that
#       we have to handle hidden rows/cols somehow.
#       There are multiple possible solutions:
#       1. Work on the csv files, by turning them into xls again. Easiest solution, requires little work
#           Problem: width and height are lost, which are required for certain metrics
#       2. Remove hidden rows/cols from the xls, by shifting data. Supported by openpyxl
#           Problem: Shifts only data, not formatting width and height are now wrongly assigned.
#       3. I rewrote the annotation preprocessor to skip over hidden rows and columns,
#           by creating a list of unhidden col_letters and unhidden row_indices, which I then can access with the
#           annotation coordinates
#           Problem: As label regions still have to be merged, it would require a major refactor
#           to enable label regions to span over hidden rows/columns.


# TODO: Investigate, what is running so long with huge files, and think to either remove or document skipping of huge files!
# TODO: Investigate, which files contribute with the extremely worse precision
# TODO: Improve logging of skipped wb & sheets, so the reasons are transparent, uniformly logged and in the same place, maybe introduce custom exceptions for that

class Dataset(object):
    def __init__(self, path, name, label_region_loader: LabelRegionLoader,
                 annotations_file_name="preprocessed_annotations_elements.json"):
        self.path = path
        self.name = name
        self.label_region_loader = label_region_loader
        self.file_size_cap = 1000 * 100  # 100 kb
        self.cap_file_size = True
        self.annotations_file_name = annotations_file_name

    @cached_property
    def _annotations(self):
        annotation_file = join(self.path, self.annotations_file_name)
        with open(annotation_file) as f:
            data = f.read()
        return json.loads(data)

    def _get_sheet_annotations(self, xls_name, sheetname):
        annotations_key = xls_name + '_' + sheetname + '.csv'
        return self._annotations[annotations_key]

    def _get_xls_file_paths(self, exceptions: List[str] = None):
        xls_file_directory = join(self.path, "xls")
        xls_files = sorted(
            [xls_file for xls_file in listdir(xls_file_directory) if isfile(join(xls_file_directory, xls_file))])

        if exceptions is not None:
            xls_files = [f for f in xls_files if f not in exceptions]

        return [
            join(xls_file_directory, xls_file)
            for xls_file in xls_files
            if self.cap_file_size is False or getsize(join(xls_file_directory, xls_file)) < self.file_size_cap
        ]

    @staticmethod
    def worksheet_contains_hidden(worksheet: Worksheet):
        hidden_cols = [dim for _, dim in worksheet.column_dimensions.items() if dim.hidden is True]
        hidden_rows = [dim for _, dim in worksheet.row_dimensions.items() if dim.hidden is True]
        if len(hidden_cols) > 0 or len(hidden_rows) > 0:
            # Contains hidden
            return True
        return False

    @staticmethod
    def _get_workbooks(xls_file_paths: List[str]):
        """Generator for all Workbook Objects"""
        for i, xls_path in enumerate(xls_file_paths):
            try:
                wb = load_workbook(xls_path)
            except:
                # Something went wrong with loading the workbook, log and skip
                logger.warning(f"Could not load workbook {xls_path}")
                continue
            wb.path = xls_path  # Path is wrongly defaulted to /xl/workbook.xml
            logger.debug(f"Loading workbook {i}/{len(xls_file_paths)}: {xls_path}")
            yield wb

    def get_sheet_data(self, exceptions: List[str] = None) -> Generator[SheetData, None, None]:
        """Generator for all Sheet Data Objects"""
        for workbook in Dataset._get_workbooks(self._get_xls_file_paths(exceptions)):
            for sheet in workbook:
                if Dataset.worksheet_contains_hidden(sheet):
                    logger.debug(
                        f"The sheet {sheet.title} of workbook {workbook.path} contains hidden cols or rows, which we cant handle yet!")
                    continue
                wb_path = sheet.parent.path
                xls_file_name = split(wb_path)[1]
                try:
                    sheet_annotations = self._get_sheet_annotations(xls_file_name, sheet.title)
                except KeyError:
                    # No annotation for this sheet exists (probably empty or a graph)
                    logger.debug(f"\tNo Annotation found for {sheet.title} of {workbook.path}")
                    continue

                logger.info(f"\tLoading sheet {sheet.title} of {workbook.path}")
                label_regions, table_definitions = self.label_region_loader.preprocess_annotations(
                    sheet,
                    sheet_annotations,
                )
                yield SheetData(sheet, label_regions, table_definitions)

    def get_specific_sheetdata(self, xls_file: str, sheet_name: str) -> SheetData:
        xls_file_path = join(self.path, "xls", xls_file)
        wb = load_workbook(xls_file_path)

        wb.path = xls_file_path  # Path is wrongly defaulted to /xl/workbook.xml
        sheet = wb[sheet_name]
        if Dataset.worksheet_contains_hidden(sheet):
            raise NotImplementedError("This sheet contains hidden cols or rows, which we cant handle yet!")
        sheet_annotations = self._get_sheet_annotations(xls_file, sheet.title)
        label_regions, table_definitions = self.label_region_loader.preprocess_annotations(
            sheet,
            sheet_annotations,
        )
        return SheetData(sheet, label_regions, table_definitions)

    def summarize(self):
        """Summarizes Dataset Annotations"""
        data = self._annotations

        analytical_data = {}
        for sheet, sheetdata in data.items():
            analytical_data[sheet] = {
                "table_count": 0,
            }
            for region in sheetdata["regions"]:
                if region["region_type"] == "Table":
                    analytical_data[sheet]["table_count"] += 1

        print(f"Total Sheets: {len(analytical_data.keys())}")
        print(
            f"Single Table Sheets: {len([sheet for sheet, analytical_sheetdata in analytical_data.items() if analytical_sheetdata['table_count'] == 1])}")
        print(
            f"Multi Table Sheets: {len([sheet for sheet, analytical_sheetdata in analytical_data.items() if analytical_sheetdata['table_count'] > 1])}")

        print(f"Total Annotatd Tables: {sum([data['table_count'] for data in analytical_data.values()])}")

    def __str__(self):
        return f"Dataset: {self.name}"
