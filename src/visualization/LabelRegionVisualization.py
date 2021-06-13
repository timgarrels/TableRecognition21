import logging
from hashlib import md5
from typing import List

from openpyxl import Workbook
from openpyxl.styles import Color, PatternFill

from labelregions.LabelRegion import LabelRegion

logger = logging.getLogger(__name__)


def color_from_int(i: int) -> str:
    """Maps an integer to a color"""
    h = md5()
    h.update(str(i).encode())
    return h.hexdigest()[:8]


def add_lr_visualization_sheet(lrs: List[LabelRegion], workbook: Workbook, sheet_name="Label Region Visualization"):
    """Creates a colorful spreadsheet from the lr data"""
    ws = workbook.create_sheet(sheet_name)
    for i, lr in enumerate(lrs):
        color = Color(rgb=color_from_int(i))
        fill = PatternFill(patternType='solid', fgColor=color)
        # Y is our row, X our column
        for y in range(lr.top, lr.bottom + 1):
            for x in range(lr.left, lr.right + 1):
                # Cell is referred by row, col
                d = ws.cell(y, x)
                d.value = f"{i} - {str(lr.type.value)}"
                d.fill = fill
