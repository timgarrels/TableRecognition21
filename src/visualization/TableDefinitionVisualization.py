from typing import List

from openpyxl import Workbook
from openpyxl.styles import Border, Side

from labelregions.BoundingBox import BoundingBox


def overlay_table_definition_visualization(
        tfs: List[BoundingBox],
        wb: Workbook,
        sheet_name="Table Definition Visualization",
        border_color="ff00ff",
):
    """Adds bounding boxes in the form of colored borders to a given sheet. Sheet has to exist already!"""
    ws = wb[sheet_name]
    for tf in tfs:
        # Visualize each table by adding cell borders (no bulk cell border in openpyxl)
        top_cells = [(x, tf.top) for x in range(tf.left, tf.right + 1)]
        bottom_cells = [(x, tf.bottom) for x in range(tf.left, tf.right + 1)]
        left_cells = [(tf.left, y) for y in range(tf.top, tf.bottom + 1)]
        right_cells = [(tf.right, y) for y in range(tf.top, tf.bottom + 1)]

        border_style = 'thick'
        top_border = Border(top=Side(border_style=border_style, color=border_color))
        bottom_border = Border(bottom=Side(border_style=border_style, color=border_color))
        left_border = Border(left=Side(border_style=border_style, color=border_color))
        right_border = Border(right=Side(border_style=border_style, color=border_color))

        cells_to_borders = [
            (top_cells, top_border),
            (bottom_cells, bottom_border),
            (left_cells, left_border),
            (right_cells, right_border),
        ]

        for cells, border in cells_to_borders:
            for x, y in cells:
                ws.cell(y, x).border = border

        # Corners
        ws.cell(tf.top, tf.left).border = Border(top=Side(border_style=border_style, color=border_color),
                                                 left=Side(border_style=border_style, color=border_color))
        ws.cell(tf.top, tf.right).border = Border(top=Side(border_style=border_style, color=border_color),
                                                  right=Side(border_style=border_style, color=border_color))
        ws.cell(tf.bottom, tf.left).border = Border(bottom=Side(border_style=border_style, color=border_color),
                                                    left=Side(border_style=border_style, color=border_color))
        ws.cell(tf.bottom, tf.right).border = Border(bottom=Side(border_style=border_style, color=border_color),
                                                     right=Side(border_style=border_style, color=border_color))
