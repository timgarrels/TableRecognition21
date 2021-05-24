import logging
import sys
from os.path import join

from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet

from ExhaustiveSearch import ExhaustiveSearch
from GeneticSearch import GeneticSearch
from GeneticSearchConfiguration import GeneticSearchConfiguration
from LabelRegionPreprocessor import LabelRegionPreprocessor
from Rater import Rater
from SpreadSheetGraph import SpreadSheetGraph

logging.basicConfig(stream=sys.stdout, level=logging.INFO)

logger = logging.getLogger(__name__)

DATA_DIR = "data"
VISUALIZATIONS_DIR = "visualizations"
ANNOTATIONS = "annotations_elements.json"
SPREADSHEET = "andrea_ring__3__HHmonthlyavg.xlsx"
ANNOTATION_FILE = join(DATA_DIR, ANNOTATIONS)
SPREADSHEET_FILE = join(DATA_DIR, SPREADSHEET)


def main():
    logger.info("Creating Label Regions...")
    wb = load_workbook(SPREADSHEET_FILE)
    sheetname = wb.sheetnames[0]
    sheet: Worksheet = wb[sheetname]
    preprocessor = LabelRegionPreprocessor()

    label_regions = preprocessor.preproces_annotations(
        ANNOTATION_FILE,
        SPREADSHEET_FILE,
        sheetname,
    )
    LabelRegionPreprocessor.visualize_lrs(label_regions, out=join(VISUALIZATIONS_DIR, 'lrs.xlsx'))

    sheet_graph = SpreadSheetGraph.from_label_regions_and_sheet(label_regions, sheet)
    logger.debug(f"Edge Count: {len(sheet_graph.edge_list)}")
    sheet_graph.visualize(out=join(VISUALIZATIONS_DIR, 'original_graph'))

    if len(sheet_graph.nodes) <= 10:
        # Less than 11 nodes, do exhaustive search
        search = ExhaustiveSearch(
            sheet_graph,
            Rater(),
        )
    else:
        search = GeneticSearch(
            sheet_graph,
            # TODO: Rater should be trained
            Rater(),
            GeneticSearchConfiguration(rand_mut_p=0.1, cross_mut_p=0.5, n_gen=200),
        )

    fittest, fittest_rating = search.run()
    logger.debug(f"Best rating: {fittest_rating}")

    logger.info("Visualizing Fittest Graph Partition...")
    sheet_graph.edge_toggle_list = fittest
    sheet_graph.visualize(out=join(VISUALIZATIONS_DIR, 'fittest_graph'))


if __name__ == "__main__":
    main()
